import socketserver
import threading
import openpyxl as ex
import json
#   A    B       C          D             E             F      G1
# Name, IP, last connect, Admin, command to execute, number,


class TCPRequestHandler(socketserver.BaseRequestHandler):
    def handle(self):
        wb = ex.load_workbook('users.xlsx')
        ws = wb['users']

        def check(name):
            nonlocal wb, ws
            num = ws['G1'].value
            counter = 1
            is_in_list = False
            for col in ws.iter_cols(min_row=1, max_col=1):
                print(col)
                for cell in col:
                    if counter >= num:
                        break
                    if cell.value() == name:
                        is_in_list = True
                    counter += 1
            return is_in_list

        def create_client(name, ip, last_connect, admin, command):
            nonlocal wb, ws
            number_of_users = ws['G1'].value
            ws[f'A{number_of_users + 1}'].value = name
            ws[f'B{number_of_users + 1}'].value = ip
            ws[f'C{number_of_users + 1}'].value = last_connect
            ws[f'D{number_of_users + 1}'].value = admin
            ws[f'E{number_of_users + 1}'].value = command
            ws[f'F{number_of_users + 1}'].value = number_of_users + 1
            ws['G1'].value += 1

        try:
            data_res = str(self.request.recv(1024), 'utf-8')
            data = json.loads(data_res)
            if not check(data['name']):
                create_client(data['name'], 123, 123, data['admin'], 123)
            response = bytes('No commands', 'utf-8')
            self.request.sendall(response)
            wb.save('users.xlsx')
        except Exception as e:
            print(f'Excepted: {e} || in handle')


if __name__ == '__main__':
    ip = 'localhost'
    port = 27036
    server = socketserver.TCPServer((ip, port), TCPRequestHandler)
    server.wb = ex.load_workbook('users.xlsx')
    try:
        server_thread = threading.Thread(target=server.serve_forever(), daemon=True)
        server_thread.start()
    except Exception as e:
        print(f'Excepted: {e}')
